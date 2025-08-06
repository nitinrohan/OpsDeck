import datetime
import socket
import psycopg2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference

LOG_FILE = "reports/health_report.xlsx"

DB_CONFIG = {
    'dbname': 'testdb',
    'user': 'testuser',
    'password': 'your_password_here',
    'host': 'localhost',
    'port': '5432'
}

def check_network():
    try:
        socket.create_connection(("8.8.8.8", 53), timeout=2)
        return "UP"
    except Exception:
        return "DOWN"

def check_postgres():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        conn.close()
        return "REACHABLE"
    except Exception:
        return "UNREACHABLE"

def write_to_excel(timestamp, network_status, db_status):
    try:
        wb = load_workbook(LOG_FILE)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Network", "PostgreSQL"])
        header_font = Font(bold=True)
        for cell in ws["1:1"]:
            cell.font = header_font

    net_val = 1 if network_status == "UP" else 0
    db_val = 1 if db_status == "REACHABLE" else 0
    ws.append([timestamp, net_val, db_val])
    wb.save(LOG_FILE)

def add_chart():
    wb = load_workbook(LOG_FILE)
    ws = wb.active

    chart = LineChart()
    chart.title = "OpsDeck System Health Over Time"
    chart.y_axis.title = "Status (1=UP/Reachable, 0=DOWN)"
    chart.x_axis.title = "Timestamp"

    data = Reference(ws, min_col=2, max_col=3, min_row=1, max_row=ws.max_row)
    categories = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 10
    chart.width = 20

    ws.add_chart(chart, f"E3")  # Position the chart on the sheet
    wb.save(LOG_FILE)

if __name__ == "__main__":
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    network_status = check_network()
    db_status = check_postgres()
    write_to_excel(timestamp, network_status, db_status)
    add_chart()
    print("✅ Health check logged and chart updated.")
